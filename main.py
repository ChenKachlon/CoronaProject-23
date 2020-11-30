import openpyxl

# Pointers to data base
wb = openpyxl.load_workbook("Hospital_Data_Base.xlsx")
ws = wb.worksheets[2]
ws_beds = wb.worksheets[1]

def Get_Row_Indicator(max_row):
    return max_row + 1

def ConvertToStr(columns,rows):
    """Convert row num and column num to str by excel format"""
    return str(chr(columns+64) + str(rows))

def Save_Data():
    """Save details in data base"""
    wb.save("Hospital_Data_Base.xlsx")

def add_patient():
    """Add patient to the data base"""
    row_location = Get_Row_Indicator(ws.max_row)
    patient_id = input("Please enter id: ")
    degree_of_difficulty = input("Please enter the degree of difficulty (mildly ill,medium ill,seriously ill,dying): ")
    ventilators = input("Need ventilator? (1-Yes,0-No): ")
    department = input("Department? ")
    data = [patient_id,degree_of_difficulty,ventilators,department]
    for i in range(4):
        ws[ConvertToStr(i+1,row_location)] = data[i]

def release_patient():
    """Release patients from the data base"""
    row_location = Get_Row_Indicator(ws.max_row)
    release_id = input("Please enter the id of the patient you want to release: ")
    flag = 0
    for i in range(2,row_location):
        if str(ws.cell(row=i, column=1).value) == str(release_id):
            ws.delete_rows(i)
            flag = 1
    if flag == 0:
        print("Sorry, Id doesn't exists in the data base")
    else:
        print("Patient released")

def amount_of_beds_in_given_department():
    """Calculate the amount of beds in given department"""
    row_location_beds = Get_Row_Indicator(ws_beds.max_row)
    amount = 0
    department_name = input("Please enter department name: ")
    for i in range(3,row_location_beds):
        if str(ws_beds.cell(row=i,column=1).value) == str(department_name):
            amount += 1
    if amount != 0:
        print("Amount of beds in {0} department is: {1}".format(department_name, amount))
    else:
        print("Department doesn't exist")
    return amount

def amount_of_beds_in_given_hospital():
    """Calculate the amount of beds in the hospital"""
    row_location_beds = Get_Row_Indicator(ws_beds.max_row)
    amount = 0
    for i in range(3, row_location_beds):
        amount += 1
    if amount != 0:
        print("Amount of beds in the hospital is: {0}".format(amount))
    else:
        print("Zero beds!!!")
    return amount

def amount_of_patients_in_given_department():
    """Calculate the amount of patients in given department"""
    row_location = Get_Row_Indicator(ws.max_row)
    amount = 0
    department_name = input("Please enter department name: ")
    for i in range(2,row_location):
        if str(ws.cell(row=i,column=4).value) == str(department_name):
            amount += 1
    if amount != 0:
        print("Amount of patients in {0} department is: {1}".format(department_name, amount))
    else:
        print("Department doesn't exist")
    return amount

def amount_of_patients_in_given_hospital():
    """Calculate the amount of patients in the hospital """
    row_location = Get_Row_Indicator(ws.max_row)
    amount = 0
    for i in range(2, row_location):
        amount += 1
    if amount != 0:
        print("Amount of patients in the hospital is: {0}".format(amount))
    else:
        print("No patients!!!")
    return amount

def Sub_Amounts_Beds():
    """Sub Amounts"""
    return amount_of_beds_in_given_hospital() - amount_of_beds_in_given_department()

def Sub_Amounts_Patients():
    """Sub Amounts"""
    return amount_of_patients_in_given_hospital() - amount_of_patients_in_given_department()

def main():
    # add_patient()
    # release_patient()
    # amount_of_beds_in_given_department()
    # amount_of_beds_in_given_hospital()
    # amount_of_patients_in_given_department()
    print(Sub_Amounts_Patients())
    Save_Data()

main()
